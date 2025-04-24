from sqlalchemy import create_engine, text
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows


DB_CONFIG = {
    "dbname": "your_database",
    "user": "username",
    "password": "password",
    "host": "localhost",
    "port": "5432"
}

def clear_screen():
    import os
    os.system("cls" if os.name == "nt" else "clear")

def get_engine():
    return create_engine(
        f"postgresql://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['dbname']}"
    )

def fetch_meal_data(engine, start_date, end_date):
    query = f"""
    WITH meal_data AS (
        SELECT 
            m.meals_date, m.meal_type, m.perm_id,
            s.first_name, s.last_name, s.staff,
            COALESCE(
                CASE 
                    WHEN e.frm_code = 'Free' THEN '1'
                    WHEN e.frm_code = 'Reduced' THEN '2'
                    ELSE '3'
                END, '3'
            ) AS meals_code
        FROM meals m
        JOIN students s ON m.perm_id = s.perm_id
        LEFT JOIN eligibility e ON m.perm_id = e.perm_id
        WHERE m.meals_date BETWEEN '{start_date}' AND '{end_date}'
    )
    SELECT * FROM meal_data ORDER BY meals_date, meal_type, perm_id;
    """
    return pd.read_sql(query, engine)

def export_meal_sheets(df, engine, school_name, start_date):
    wb = Workbook()
    wb.remove(wb.active)

    for (meals_date, meal_type), group in df.groupby(["meals_date", "meal_type"]):
        sheet_name = f"{meals_date}_{meal_type}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        ws.append(["perm_id", "first_name", "last_name", "staff", "meals_code"])
        for row in group.itertuples(index=False):
            ws.append([row.perm_id, row.first_name, row.last_name, row.staff, row.meals_code])

        # Totals
        total_1s = sum(group["meals_code"] == '1')
        total_2s = sum(group["meals_code"] == '2')
        total_3s = sum(group["meals_code"] == '3')
        total_served = len(group)

        ws.append([])
        ws.append(["Totals"])
        ws.append(["Meals Code 1", total_1s])
        ws.append(["Meals Code 2", total_2s])
        ws.append(["Meals Code 3", total_3s])
        ws.append(["Total Served", total_served])

        insert_totals(engine, meal_type, meals_date, total_1s, total_2s, total_3s, total_served)

    filename = f"{school_name}_{start_date}_meal_report.xlsx"
    wb.save(filename)
    print(f"Report saved: {filename}")

def insert_totals(engine, meal_type, date, t1, t2, t3, total):
    table_name = f"{meal_type}_totals"
    query = text(f"""
        INSERT INTO {table_name} (meals_date, total_1s, total_2s, total_3s, total_served)
        VALUES (:date, :t1, :t2, :t3, :total)
        ON CONFLICT (meals_date) DO UPDATE SET
            total_1s = EXCLUDED.total_1s,
            total_2s = EXCLUDED.total_2s,
            total_3s = EXCLUDED.total_3s,
            total_served = EXCLUDED.total_served;
    """)
    with engine.connect() as conn:
        conn.execute(query, {"date": date, "t1": t1, "t2": t2, "t3": t3, "total": total})
        conn.commit()

def generate_summary_tables(engine, start_date, end_date):
    # Get totals and orders
    totals_query = f"""
    SELECT 
        COALESCE(b.meals_date, l.meals_date) AS order_date,
        l.total_1s AS lunch_1s, l.total_2s AS lunch_2s, l.total_3s AS lunch_3s, l.total_served AS lunch_total,
        b.total_1s AS breakfast_1s, b.total_2s AS breakfast_2s, b.total_3s AS breakfast_3s, b.total_served AS breakfast_total
    FROM lunch_totals l
    FULL OUTER JOIN breakfast_totals b ON l.meals_date = b.meals_date
    WHERE COALESCE(b.meals_date, l.meals_date) BETWEEN '{start_date}' AND '{end_date}'
    ORDER BY order_date;
    """
    orders_query = f"""
    SELECT order_date, lunch_order, breakfast_order FROM orders
    WHERE order_date BETWEEN '{start_date}' AND '{end_date}'
    ORDER BY order_date;
    """

    df_totals = pd.read_sql(totals_query, engine)
    df_orders = pd.read_sql(orders_query, engine)
    return df_orders, df_totals

def summarize_meals(df_orders, df_totals, meal_type):
    cols = ["order_date", f"{meal_type}_1s", f"{meal_type}_2s", f"{meal_type}_3s", f"{meal_type}_total"]
    df = pd.merge(df_orders[["order_date", f"{meal_type}_order"]], df_totals[cols], on="order_date", how="left")

    df["leftover"] = df[f"{meal_type}_order"] - df[f"{meal_type}_total"]
    df["FRM_total"] = df[f"{meal_type}_1s"] + df[f"{meal_type}_2s"] + df[f"{meal_type}_3s"]

    return df.rename(columns={
        f"{meal_type}_order": "meal_order",
        f"{meal_type}_total": "total_served",
        f"{meal_type}_1s": "total_1s",
        f"{meal_type}_2s": "total_2s",
        f"{meal_type}_3s": "total_3s",
    })[["order_date", "meal_order", "total_served", "leftover", "total_1s", "total_2s", "total_3s", "FRM_total"]]


def write_summary_workbook(lunch_summary_df, breakfast_summary_df, filename=None, school=None, month_report=None):
    if filename is None and school and month_report:
        filename = f"{school}_{month_report}_summary.xlsx"
    elif filename is None:
        filename = "summary.xlsx"

    wb = Workbook()

    def write_sheet(ws, df, title):
        ws.title = title
        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        # Add Totals Row
        totals = ["Totals"]
        for col in df.columns[1:]:
            if pd.api.types.is_numeric_dtype(df[col]):
                totals.append(df[col].sum())
            else:
                totals.append("")
        ws.append(totals)
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    # Write Lunch Summary
    ws_lunch = wb.active
    write_sheet(ws_lunch, lunch_summary_df, "Lunch Summary")

    # Write Breakfast Summary
    ws_breakfast = wb.create_sheet()
    write_sheet(ws_breakfast, breakfast_summary_df, "Breakfast Summary")

    wb.save(filename)
    print(f"Summary workbook saved as {filename}")


# Optional CLI or menu integration
def main():
    clear_screen()
    engine = get_engine()
    school = input("Enter school name: ")
    month_report = input("Enter the report month: ")
    start_date = input("Enter start date (YYYY-MM-DD): ")
    end_date = input("Enter end date (YYYY-MM-DD): ")

    df = fetch_meal_data(engine, start_date, end_date)
    export_meal_sheets(df, engine, school, start_date)

    df_orders, df_totals = generate_summary_tables(engine, start_date, end_date)
    lunch_df = summarize_meals(df_orders, df_totals, "lunch")
    breakfast_df = summarize_meals(df_orders, df_totals, "breakfast")

    print("\nLunch Summary:\n", lunch_df)
    print("\nBreakfast Summary:\n", breakfast_df)
    write_summary_workbook(lunch_df, breakfast_df, school=school, month_report=month_report)


if __name__ == "__main__":
    main()
